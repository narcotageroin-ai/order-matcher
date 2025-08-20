
import os
from io import BytesIO
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Order Matcher", page_icon="📦", layout="centered")
st.title("📦 Order Matcher — универсальная версия (фикс)")

# -------------------- УТИЛИТЫ --------------------
def normalize_key(x):
    if pd.isna(x):
        return None
    return str(x).strip()

def guess_qty_column(columns):
    patterns = ['кол-во','количество','qty','quantity','qnt','шт']
    low = [str(c).strip().lower() for c in columns]
    for p in patterns:
        for i, name in enumerate(low):
            if p in name:
                return columns[i]
    return columns[0] if len(columns) else None

def read_any(uploaded_file):
    """Читаем CSV/XLS/XLSX корректным способом и, для XLSX, даём Workbook для сохранения формата."""
    name = uploaded_file.name
    ext = os.path.splitext(name)[1].lower()
    raw = uploaded_file.getvalue()

    if ext == ".csv":
        df = pd.read_csv(BytesIO(raw))
        return df, "csv", raw, None

    if ext == ".xls":
        # старый формат Excel — читаем через xlrd; формат сохранить не получится
        df = pd.read_excel(BytesIO(raw), engine="xlrd")
        return df, "xls", raw, None

    if ext in [".xlsx", ".xlsm"]:
        df = pd.read_excel(BytesIO(raw), engine="openpyxl")
        wb = load_workbook(BytesIO(raw), keep_vba=(ext==".xlsm"))
        return df, "xlsx", raw, wb

    st.error(f"❌ Неподдерживаемый формат: {ext}")
    st.stop()

def detect_header_row(ws, pandas_headers, max_scan=30):
    """Ищем строку заголовка на листе XLSX, сравнивая с именами колонок из pandas."""
    good_names = set([str(h) for h in pandas_headers if not str(h).startswith('Unnamed:')])
    best_r, best_hits = 1, -1
    for r in range(1, min(ws.max_row, max_scan)+1):
        vals = [str(c.value) if c.value is not None else '' for c in ws[r]]
        hits = sum(1 for v in vals if v in good_names and v != '')
        if hits > best_hits:
            best_hits, best_r = hits, r
    return best_r

def build_header_map(ws, header_row):
    m = {}
    row = ws[header_row]
    for j, cell in enumerate(row, start=1):
        if cell.value is not None:
            m[str(cell.value)] = j
    return m

# -------------------- UI: загрузка --------------------
file_order = st.file_uploader("Загрузите файл заказа (CSV / XLS / XLSX)", type=["csv","xls","xlsx","xlsm"], key="order")
file_supplier = st.file_uploader("Загрузите файл поставщика (CSV / XLS / XLSX)", type=["csv","xls","xlsx","xlsm"], key="supplier")

if not (file_order and file_supplier):
    st.info("Загрузите оба файла, чтобы продолжить.")
    st.stop()

df_order, order_type, order_bytes, order_wb = read_any(file_order)
df_supplier, supplier_type, supplier_bytes, supplier_wb = read_any(file_supplier)

st.subheader("Наш файл заказа (первые строки)")
st.dataframe(df_order.head(10))
st.subheader("Файл поставщика (первые строки)")
st.dataframe(df_supplier.head(10))

# -------------------- UI: сопоставление --------------------
st.markdown("### Настройка сопоставления")
st.caption("Выберите ключевые поля **в порядке приоритета** (сначала Артикул, затем Штрихкод и т.д.)")

order_keys = st.multiselect("🔑 Поля поиска (наш файл)", df_order.columns)
supplier_keys = st.multiselect("🔍 Поля поиска (файл поставщика)", df_supplier.columns)

if len(order_keys) != len(supplier_keys):
    st.warning("Выберите одинаковое количество ключей в обоих файлах.")
    st.stop()

qty_col_order = st.selectbox("📦 Колонка количества (наш файл)", df_order.columns, index=max(0, list(df_order.columns).index(guess_qty_column(df_order.columns)) if guess_qty_column(df_order.columns) in df_order.columns else 0))

# Варианты задания колонки количества у поставщика
st.markdown("**Куда записывать количество у поставщика?**")
supplier_qty_mode = st.radio("Способ выбора", ["Выбрать существующую колонку", "Создать новую колонку"], horizontal=True)

new_qty_name = None
if supplier_qty_mode == "Выбрать существующую колонку":
    supplier_qty_col = st.selectbox("✏️ Колонка количества у поставщика", df_supplier.columns)
else:
    default_name = "Количество"
    new_qty_name = st.text_input("Название новой колонки", value=default_name)
    supplier_qty_col = new_qty_name

overwrite = st.checkbox("Перезаписывать уже заполненные значения у поставщика", value=False)

# Доп. опции
case_insensitive = st.checkbox("Игнорировать регистр при сравнении (A==a)", value=True)
trim_spaces = st.checkbox("Игнорировать пробелы по краям", value=True)

# -------------------- ПОДГОТОВКА СЛОВАРЕЙ --------------------
def key_norm(x):
    s = normalize_key(x)
    if s is None:
        return None
    if trim_spaces:
        s = s.strip()
    if case_insensitive:
        s = s.lower()
    return s

dicts = []
for okey in order_keys:
    d = {}
    for a, q in zip(df_order[okey], df_order[qty_col_order]):
        na = key_norm(a)
        if na is not None:
            d[na] = q
    dicts.append(d)

# -------------------- ОБРАБОТКА --------------------
if st.button("🚀 Сформировать файл"):
    stats = {}
    matched = 0
    total = len(df_supplier)

    # ===== Сценарий XLSX: сохраняем формат через openpyxl =====
    if supplier_type == "xlsx":
        ws = supplier_wb.active
        header_row = detect_header_row(ws, df_supplier.columns)
        col_map = build_header_map(ws, header_row)
        pos_map = {str(c): i+1 for i, c in enumerate(df_supplier.columns)}  # позиция по pandas

        # qty-колонка
        if supplier_qty_col not in df_supplier.columns and supplier_qty_mode == "Создать новую колонку":
            # создадим в конце таблицы (только в pandas-файле для расчёта индекса)
            df_supplier[supplier_qty_col] = None
            pos_map = {str(c): i+1 for i, c in enumerate(df_supplier.columns)}
            # в самом Excel — добавим новый столбец справа
            insert_idx = ws.max_column + 1
            ws.cell(row=header_row, column=insert_idx).value = supplier_qty_col
            qty_idx = insert_idx
        else:
            qty_idx = col_map.get(str(supplier_qty_col))
            if qty_idx is None:
                # fallback по позиции (для Unnamed)
                qty_idx = pos_map.get(str(supplier_qty_col))

        key_idxs = []
        fallback_used = []
        for skey in supplier_keys:
            idx = col_map.get(str(skey))
            if idx is None:
                idx = pos_map.get(str(skey))
                if idx is not None:
                    fallback_used.append(str(skey))
            key_idxs.append(idx)

        if any(i is None for i in [qty_idx] + key_idxs):
            st.error("❌ Не удалось сопоставить некоторые столбцы (даже по позиции). Проверьте выбор.")
            st.stop()

        if fallback_used:
            st.warning("⚠️ Столбцы без названия сопоставлены по позиции: " + ", ".join(fallback_used))

        start_row = header_row + 1
        for r in range(start_row, ws.max_row + 1):
            # если нельзя перезаписывать — пропускаем непустые
            if not overwrite and ws.cell(row=r, column=qty_idx).value is not None:
                continue

            placed = False
            for d_idx, k_idx in enumerate(key_idxs, start=1):
                val = key_norm(ws.cell(row=r, column=k_idx).value)
                if val is None:
                    continue
                q = dicts[d_idx - 1].get(val)
                if q is not None:
                    ws.cell(row=r, column=qty_idx).value = q
                    key_name = f"{order_keys[d_idx-1]} -> {supplier_keys[d_idx-1]}"
                    stats[key_name] = stats.get(key_name, 0) + 1
                    matched += 1
                    placed = True
                    break

        not_found = total - matched

        st.subheader("📊 Результаты")
        st.write(f"Всего строк у поставщика: **{total}**")
        for k, v in stats.items():
            st.write(f"🔍 Найдено по {k}: **{v}**")
        st.write(f"⚠️ Не найдено: **{not_found}**")

        out = BytesIO()
        supplier_wb.save(out)
        st.success("✅ Готово! Скачайте результат.")
        st.download_button("⬇ Скачать XLSX (с сохранением оформления)",
            data=out.getvalue(),
            file_name="supplier_with_qty.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ===== CSV / XLS: через pandas =====
    else:
        df = df_supplier.copy()

        if supplier_qty_col not in df.columns and supplier_qty_mode == "Создать новую колонку":
            df[supplier_qty_col] = None

        if supplier_qty_col not in df.columns:
            st.error("❌ Укажите существующую колонку или создайте новую.")
            st.stop()

        # маска для незаполненных (если не перезаписываем)
        if overwrite:
            editable_mask = pd.Series([True] * len(df))
        else:
            editable_mask = df[supplier_qty_col].isna() if supplier_qty_col in df.columns else pd.Series([True]*len(df))

        # Последовательно применяем ключи (приоритетно)
        for d_idx, (okey, skey) in enumerate(zip(order_keys, supplier_keys), start=1):
            d = dicts[d_idx - 1]
            # нормализуем значения supplier для сопоставления
            supplier_norm = df[skey].map(key_norm)
            before = df.loc[editable_mask, supplier_qty_col].notna().sum()
            df.loc[editable_mask, supplier_qty_col] = df.loc[editable_mask, skey].map(lambda x: d.get(key_norm(x)))
            after = df.loc[editable_mask, supplier_qty_col].notna().sum()
            stats[f"{okey} -> {skey}"] = int(after - before)

        matched = int(df[supplier_qty_col].notna().sum())
        not_found = int(len(df) - matched)

        st.subheader("📊 Результаты")
        st.write(f"Всего строк у поставщика: **{len(df)}**")
        for k, v in stats.items():
            st.write(f"🔍 Найдено по {k}: **{v}**")
        st.write(f"⚠️ Не найдено: **{not_found}**")

        out = BytesIO()
        if supplier_type == "csv":
            df.to_csv(out, index=False, encoding="utf-8-sig")
            fname = "supplier_with_qty.csv"
            mime = "text/csv"
        else:
            # .xls сохраняем как .xlsx (надёжнее)
            df.to_excel(out, index=False, engine="openpyxl")
            fname = "supplier_with_qty.xlsx"
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        st.success("✅ Готово! Скачайте результат.")
        st.download_button("⬇ Скачать результат", data=out.getvalue(), file_name=fname, mime=mime)
